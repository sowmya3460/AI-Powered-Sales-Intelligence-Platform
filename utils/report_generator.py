import io
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def generate_excel_report(dfs: dict) -> bytes:
    """dfs = {'Sheet Name': dataframe, ...}"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            ws = writer.sheets[sheet_name[:31]]
            # Style header
            header_fill = PatternFill("solid", fgColor="B5DEFF")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="2D2D3A")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    return output.getvalue()


def generate_pdf_report(title: str, sections: list) -> bytes:
    """sections = [{'heading': str, 'data': dataframe or str}]"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                  fontSize=20, textColor=colors.HexColor("#7C6FF7"),
                                  spaceAfter=14)
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"],
                                    fontSize=13, textColor=colors.HexColor("#2D2D3A"),
                                    spaceAfter=6)
    body_style = styles["BodyText"]
    story = [Paragraph(title, title_style), Spacer(1, 12)]

    for sec in sections:
        story.append(Paragraph(sec["heading"], heading_style))
        data = sec["data"]
        if isinstance(data, pd.DataFrame):
            df_small = data.head(20)
            table_data = [list(df_small.columns)] + df_small.values.tolist()
            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B5DEFF")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#2D2D3A")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#EEF0FF")]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)
        elif isinstance(data, str):
            story.append(Paragraph(data, body_style))
        story.append(Spacer(1, 12))

    doc.build(story)
    return output.getvalue()